import frappe
from frappe import _
import json


def _assert_coordenador():
    user = frappe.session.user
    if user == "Guest":
        frappe.throw(_("Não autenticado"), frappe.AuthenticationError)
    roles = frappe.get_roles(user)
    if "System Manager" not in roles and "Coordenador Catequese" not in roles:
        frappe.throw(_("Sem permissão"), frappe.PermissionError)


@frappe.whitelist()
def get_anos_lectivos():
    _assert_coordenador()
    try:
        rows = frappe.db.sql(
            "SELECT name FROM `tabAno Lectivo` ORDER BY name DESC LIMIT 10",
            as_dict=True,
        )
        return [r.name for r in rows]
    except Exception:
        return []


@frappe.whitelist()
def get_ano_lectivo_atual():
    _assert_coordenador()
    try:
        ano = frappe.db.get_value("Ano Lectivo", {"is_current": 1}, "name")
        if not ano:
            rows = frappe.db.sql(
                "SELECT name FROM `tabAno Lectivo` ORDER BY name DESC LIMIT 1",
                as_dict=True,
            )
            ano = rows[0].name if rows else None
        return ano
    except Exception:
        return None


@frappe.whitelist()
def get_fases():
    _assert_coordenador()
    try:
        rows = frappe.db.sql(
            "SELECT name FROM `tabFase` ORDER BY name ASC",
            as_dict=True,
        )
        return [r.name for r in rows]
    except Exception:
        return []


@frappe.whitelist()
def get_letter_heads():
    _assert_coordenador()
    try:
        rows = frappe.db.sql(
            "SELECT name, content, footer FROM `tabLetter Head` WHERE disabled = 0 ORDER BY is_default DESC, name ASC",
            as_dict=True,
        )
        return rows
    except Exception:
        return []


@frappe.whitelist()
def get_retiros(ano_lectivo):
    _assert_coordenador()
    rows = frappe.db.sql("""
        SELECT
            name, titulo, data, estado,
            local, orador, tema,
            fase_1, fase_2, valor_de_contribuicao
        FROM `tabPlano de Retiro`
        WHERE ano_lectivo = %s
        ORDER BY data IS NULL ASC, data ASC
    """, (ano_lectivo,), as_dict=True)
    return rows


@frappe.whitelist()
def create_retiro(data_json):
    _assert_coordenador()
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    doc = frappe.new_doc("Plano de Retiro")
    doc.titulo      = data.get("titulo")
    doc.data        = data.get("data") or None
    doc.ano_lectivo = data.get("ano_lectivo")
    doc.estado      = data.get("estado") or "Planeado"
    doc.local       = data.get("local") or None
    doc.orador      = data.get("orador") or None
    doc.tema        = data.get("tema") or None
    doc.fase_1                 = data.get("fase_1") or None
    doc.fase_2                 = data.get("fase_2") or None
    doc.valor_de_contribuicao  = data.get("valor_de_contribuicao") or None
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    return _fetch_retiro(doc.name)


@frappe.whitelist()
def update_retiro(name, data_json):
    _assert_coordenador()
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    doc = frappe.get_doc("Plano de Retiro", name)
    doc.titulo  = data.get("titulo", doc.titulo)
    doc.data    = data.get("data") or None
    doc.estado  = data.get("estado", doc.estado)
    doc.local   = data.get("local") or None
    doc.orador  = data.get("orador") or None
    doc.tema    = data.get("tema") or None
    doc.fase_1                = data.get("fase_1") or None
    doc.fase_2                = data.get("fase_2") or None
    doc.valor_de_contribuicao = data.get("valor_de_contribuicao") or None
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    return _fetch_retiro(name)


@frappe.whitelist()
def delete_retiro(name):
    _assert_coordenador()
    if not frappe.db.exists("Plano de Retiro", name):
        frappe.throw(_("Retiro não encontrado"))
    frappe.delete_doc("Plano de Retiro", name, ignore_permissions=True)
    frappe.db.commit()
    return {"success": True}


@frappe.whitelist()
def update_estado(name, estado):
    _assert_coordenador()
    allowed = {"Planeado", "Realizado", "Cancelado"}
    if estado not in allowed:
        frappe.throw(_("Estado inválido"))
    frappe.db.set_value("Plano de Retiro", name, "estado", estado)
    frappe.db.commit()
    return {"success": True, "estado": estado}


@frappe.whitelist()
def get_programa(retiro_name):
    _assert_coordenador()
    if not frappe.db.exists("Plano de Retiro", retiro_name):
        frappe.throw(_("Retiro não encontrado"))
    rows = frappe.db.sql("""
        SELECT name, hora, actividade, responsavel, notas, idx
        FROM `tabRetiro Item`
        WHERE parent = %s
        ORDER BY idx ASC
    """, (retiro_name,), as_dict=True)
    return rows


@frappe.whitelist()
def save_programa(retiro_name, items_json):
    _assert_coordenador()
    items = json.loads(items_json) if isinstance(items_json, str) else items_json
    doc = frappe.get_doc("Plano de Retiro", retiro_name)
    doc.set("programa", [])
    for item in items:
        doc.append("programa", {
            "hora":        item.get("hora") or None,
            "actividade":  item.get("actividade") or "",
            "responsavel": item.get("responsavel") or None,
            "notas":       item.get("notas") or None,
        })
    doc.save(ignore_permissions=True)
    frappe.db.commit()
    return frappe.db.sql("""
        SELECT name, hora, actividade, responsavel, notas, idx
        FROM `tabRetiro Item`
        WHERE parent = %s ORDER BY idx ASC
    """, (retiro_name,), as_dict=True)


@frappe.whitelist()
def export_retiros(ano_lectivo, estado="", fase="", search="", fields_json=""):
    _assert_coordenador()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import date as _date
    except ImportError:
        frappe.throw(_("openpyxl não está instalado. Execute: pip install openpyxl"))

    conditions = ["ano_lectivo = %s"]
    params     = [ano_lectivo]

    if estado:
        conditions.append("estado = %s")
        params.append(estado)

    if fase:
        conditions.append("(fase_1 = %s OR fase_2 = %s)")
        params.extend([fase, fase])

    search = (search or "").strip()
    if search:
        sq = f"%{search}%"
        conditions.append("(titulo LIKE %s OR orador LIKE %s OR local LIKE %s OR tema LIKE %s)")
        params.extend([sq, sq, sq, sq])

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT name, titulo, data, estado, local, orador, tema,
               fase_1, fase_2, valor_de_contribuicao, notas
        FROM `tabPlano de Retiro`
        WHERE {where}
        ORDER BY data IS NULL ASC, data ASC
    """, params, as_dict=True)

    MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

    def fmt_date(d):
        if not d:
            return "—"
        s = str(d)[:10]
        y, m, day = s.split("-")
        return f"{int(day)} {MESES[int(m)-1]} {y}"

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _side(style="thin", color="D1D5DB"):
        return Side(style=style, color=color)

    def _border():
        s = _side()
        return Border(left=s, right=s, top=s, bottom=s)

    TITLE_FONT  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    META_FONT   = Font(name="Calibri", size=9,  color="6B7280", italic=True)
    HEADER_FONT = Font(name="Calibri", bold=True, size=9, color="374151")
    BODY_FONT   = Font(name="Calibri", size=9,  color="1F2937")
    MUTED_FONT  = Font(name="Calibri", size=9,  color="6B7280")

    TITLE_FILL  = _fill("9A7020")
    META_FILL   = _fill("FEF9EC")
    HEADER_FILL = _fill("F5DFA0")

    STATUS_DATA = {
        "Planeado":  (_fill("DBEAFE"), Font(name="Calibri", size=9, bold=True, color="1D4ED8")),
        "Realizado": (_fill("DCFCE7"), Font(name="Calibri", size=9, bold=True, color="166534")),
        "Cancelado": (_fill("FEE2E2"), Font(name="Calibri", size=9, bold=True, color="991B1B")),
    }

    # Parse selected fields (default all enabled)
    try:
        fields = json.loads(fields_json) if fields_json else {}
    except Exception:
        fields = {}

    FIELD_META = [
        ("titulo",       "Título",            35, "text"),
        ("orador",       "Orador",            20, "text"),
        ("fases",        "Fases",             15, "text"),
        ("data",         "Data",              15, "text"),
        ("local",        "Local",             20, "text"),
        ("contribuicao", "Contribuição (MZN)", 18, "num"),
        ("estado",       "Estado",            12, "status"),
        ("tema",         "Tema",              25, "text"),
        ("notas",        "Notas",             35, "text"),
    ]
    active_cols = [(k, lbl, w, ct) for k, lbl, w, ct in FIELD_META if fields.get(k, True)]
    COLS   = ["N"] + [lbl for _, lbl, _, _ in active_cols]
    WIDTHS = [5]   + [w   for _, _,   w, _ in active_cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Plano de Retiros"

    nc = len(COLS)

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    tc = ws["A1"]
    tc.value = f"Plano de Retiros — {ano_lectivo}"
    tc.font  = TITLE_FONT
    tc.fill  = TITLE_FILL
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 26

    # Meta row
    ws.merge_cells(f"A2:{get_column_letter(nc)}2")
    mc = ws["A2"]
    filters_desc = []
    if estado: filters_desc.append(f"Estado: {estado}")
    if fase:   filters_desc.append(f"Fase: {fase}")
    if search: filters_desc.append(f'Pesquisa: "{search}"')
    meta_text = f"Total: {len(rows)} retiro(s)"
    if filters_desc:
        meta_text += "  |  Filtros: " + ", ".join(filters_desc)
    meta_text += f"  |  Exportado em {_date.today().strftime('%d/%m/%Y')}"
    mc.value     = meta_text
    mc.font      = Font(name="Calibri", size=9, color="7A5A18", italic=True)
    mc.fill      = META_FILL
    mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 4

    # Header row
    HR = 4
    for ci, (col_name, width) in enumerate(zip(COLS, WIDTHS), 1):
        cell = ws.cell(row=HR, column=ci)
        cell.value = col_name
        cell.font  = Font(name="Calibri", bold=True, size=9, color="7A5A18")
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if ci == 1 else "left", vertical="center")
        cell.border = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[HR].height = 18
    ws.freeze_panes = ws.cell(row=HR + 1, column=1)

    # Data rows
    for ri, row in enumerate(rows, 1):
        rn  = HR + ri
        alt = ri % 2 == 0
        fases_str    = " + ".join(filter(None, [row.fase_1, row.fase_2]))
        contribuicao = f"{float(row.valor_de_contribuicao):.2f}" if row.valor_de_contribuicao else "—"
        s_fill, s_font = STATUS_DATA.get(row.estado, (_fill("F9FAFB"), BODY_FONT))

        FIELD_VALUES = {
            "titulo":       row.titulo or "",
            "orador":       row.orador or "",
            "fases":        fases_str,
            "data":         fmt_date(row.data),
            "local":        row.local  or "",
            "contribuicao": contribuicao,
            "estado":       row.estado or "",
            "tema":         row.tema   or "",
            "notas":        row.notas  or "",
        }

        # N column
        n_cell = ws.cell(row=rn, column=1)
        n_cell.value     = ri
        n_cell.font      = MUTED_FONT
        n_cell.alignment = Alignment(horizontal="center", vertical="center")
        n_cell.border    = Border(bottom=Side(style="thin", color="E5E7EB"))
        if alt: n_cell.fill = _fill("FFFDF5")

        for ci, (key, _, _, ctype) in enumerate(active_cols, 2):
            cell = ws.cell(row=rn, column=ci)
            cell.value  = FIELD_VALUES[key]
            cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))
            if ctype == "status":
                cell.font      = s_font
                cell.fill      = s_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ctype == "num":
                cell.font      = MUTED_FONT
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if alt: cell.fill = _fill("FFFDF5")
            else:
                cell.font = BODY_FONT
                if alt: cell.fill = _fill("FFFDF5")
        ws.row_dimensions[rn].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"Plano_Retiros_{ano_lectivo.replace('/', '-')}.xlsx"
    frappe.response.filename    = fname
    frappe.response.filecontent = buf.read()
    frappe.response.type        = "download"


def _fetch_retiro(name):
    rows = frappe.db.sql("""
        SELECT name, titulo, data, estado,
               local, orador, tema,
               fase_1, fase_2, valor_de_contribuicao
        FROM `tabPlano de Retiro`
        WHERE name = %s
    """, (name,), as_dict=True)
    return rows[0] if rows else {"name": name}
