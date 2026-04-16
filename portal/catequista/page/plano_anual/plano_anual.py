import frappe
from frappe import _
import json


def _assert_coordenador():
    user = frappe.session.user
    if user == "Guest":
        frappe.throw(_("Não autenticado"), frappe.AuthenticationError)
    roles = frappe.get_roles(user)
    if "System Manager" not in roles and "Coordenador Catequese" not in roles:
        frappe.throw(_("Sem permissão"), frappe.PermissionError)


@frappe.whitelist()
def get_actividades(ano_lectivo):
    _assert_coordenador()

    rows = frappe.db.sql("""
        SELECT
            a.name, a.actividade, a.data, a.data_original,
            a.orador, a.local, a.orcamento,
            a.tipologia, a.estado, a.notas_execucao,
            a.ano_lectivo,
            t.cor AS tipologia_cor,
            t.icone AS tipologia_icone
        FROM `tabActividade do Plano` a
        LEFT JOIN `tabTipologia Actividade` t ON t.name = a.tipologia
        WHERE a.ano_lectivo = %s
        ORDER BY a.data IS NULL ASC, a.data ASC, a.name ASC
    """, (ano_lectivo,), as_dict=True)

    return rows


@frappe.whitelist()
def get_tipologias():
    _assert_coordenador()
    return frappe.db.sql(
        "SELECT name, cor, icone FROM `tabTipologia Actividade` ORDER BY name ASC",
        as_dict=True,
    )


@frappe.whitelist()
def create_tipologia(nome, cor="", icone=""):
    _assert_coordenador()
    nome = (nome or "").strip()
    if not nome:
        frappe.throw(_("O nome da tipologia é obrigatório"))
    if frappe.db.exists("Tipologia Actividade", nome):
        frappe.throw(_("Já existe uma tipologia com esse nome"))
    doc = frappe.new_doc("Tipologia Actividade")
    doc.nome  = nome
    doc.cor   = cor.strip()  if cor   else ""
    doc.icone = icone.strip() if icone else ""
    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return {"name": doc.name, "cor": doc.cor, "icone": doc.icone}


@frappe.whitelist()
def get_anos_lectivos():
    _assert_coordenador()
    try:
        anos = frappe.db.sql(
            "SELECT name FROM `tabAno Lectivo` ORDER BY name DESC LIMIT 10",
            as_dict=True,
        )
        return [a.name for a in anos]
    except Exception:
        return []


@frappe.whitelist()
def get_ano_lectivo_atual():
    _assert_coordenador()
    try:
        # Try is_current flag first (common Frappe pattern)
        ano = None
        try:
            ano = frappe.db.get_value("Ano Lectivo", {"is_current": 1}, "name")
        except Exception:
            pass
        if not ano:
            # fallback: most recent by name (assumes "YYYY-YYYY" format)
            rows = frappe.db.sql(
                "SELECT name FROM `tabAno Lectivo` ORDER BY name DESC LIMIT 1",
                as_dict=True,
            )
            ano = rows[0].name if rows else None
        return ano
    except Exception:
        return None


@frappe.whitelist()
def create_actividade(data_json):
    _assert_coordenador()
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    doc = frappe.new_doc("Actividade do Plano")
    doc.actividade    = data.get("actividade")
    doc.tipologia     = data.get("tipologia") or None
    doc.estado        = data.get("estado") or "Pendente"
    doc.ano_lectivo   = data.get("ano_lectivo")
    doc.data          = data.get("data") or None
    doc.orador        = data.get("orador") or None
    doc.local         = data.get("local") or None
    doc.orcamento     = data.get("orcamento") or None
    doc.notas_execucao = data.get("notas_execucao") or None
    doc.insert(ignore_permissions=True)
    frappe.db.commit()

    # Return full row with tipologia details
    row = frappe.db.sql("""
        SELECT a.name, a.actividade, a.data, a.data_original,
               a.orador, a.local, a.orcamento,
               a.tipologia, a.estado, a.notas_execucao, a.ano_lectivo,
               t.cor AS tipologia_cor, t.icone AS tipologia_icone
        FROM `tabActividade do Plano` a
        LEFT JOIN `tabTipologia Actividade` t ON t.name = a.tipologia
        WHERE a.name = %s
    """, (doc.name,), as_dict=True)
    return row[0] if row else {"name": doc.name}


@frappe.whitelist()
def update_actividade(name, data_json):
    _assert_coordenador()
    data = json.loads(data_json) if isinstance(data_json, str) else data_json

    doc = frappe.get_doc("Actividade do Plano", name)

    # Preserve original date if date changes
    new_data = data.get("data") or None
    if new_data and doc.data and str(doc.data) != str(new_data) and not doc.data_original:
        doc.data_original = doc.data

    doc.actividade     = data.get("actividade", doc.actividade)
    doc.tipologia      = data.get("tipologia") or None
    doc.estado         = data.get("estado", doc.estado)
    doc.data           = new_data
    doc.orador         = data.get("orador") or None
    doc.local          = data.get("local") or None
    doc.orcamento      = data.get("orcamento") or None
    doc.notas_execucao = data.get("notas_execucao") or None
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    row = frappe.db.sql("""
        SELECT a.name, a.actividade, a.data, a.data_original,
               a.orador, a.local, a.orcamento,
               a.tipologia, a.estado, a.notas_execucao, a.ano_lectivo,
               t.cor AS tipologia_cor, t.icone AS tipologia_icone
        FROM `tabActividade do Plano` a
        LEFT JOIN `tabTipologia Actividade` t ON t.name = a.tipologia
        WHERE a.name = %s
    """, (name,), as_dict=True)
    return row[0] if row else {"name": name}


@frappe.whitelist()
def delete_actividade(name):
    _assert_coordenador()
    if not frappe.db.exists("Actividade do Plano", name):
        frappe.throw(_("Actividade não encontrada"))
    frappe.delete_doc("Actividade do Plano", name, ignore_permissions=True)
    frappe.db.commit()
    return {"success": True}


@frappe.whitelist()
def update_estado(name, estado):
    _assert_coordenador()
    allowed = {"Pendente", "Em Progresso", "Realizada", "Cancelada", "Adiada"}
    if estado not in allowed:
        frappe.throw(_("Estado inválido"))
    frappe.db.set_value("Actividade do Plano", name, "estado", estado)
    frappe.db.commit()
    return {"success": True, "estado": estado}


@frappe.whitelist()
def export_actividades(ano_lectivo, estado="", tipologias_json="", month=""):
    """
    Exports the activities plan to a styled .xlsx file.
    Respects the same filters the UI has active.
    """
    _assert_coordenador()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from collections import OrderedDict
        from datetime import date as _date
    except ImportError:
        frappe.throw(_("openpyxl não está instalado. Execute: pip install openpyxl"))

    # ── Build query filters ────────────────────────────────────────────────
    conditions = ["a.ano_lectivo = %s"]
    params     = [ano_lectivo]

    if estado:
        conditions.append("a.estado = %s")
        params.append(estado)

    tip_list = json.loads(tipologias_json) if tipologias_json else []
    if tip_list:
        placeholders = ",".join(["%s"] * len(tip_list))
        conditions.append(f"a.tipologia IN ({placeholders})")
        params.extend(tip_list)

    if month:
        conditions.append("DATE_FORMAT(a.data, '%%Y-%%m') = %s")
        params.append(month)

    where = " AND ".join(conditions)

    rows = frappe.db.sql(f"""
        SELECT
            a.name, a.actividade, a.data, a.data_original,
            a.orador, a.local, a.orcamento,
            a.tipologia, a.estado, a.notas_execucao
        FROM `tabActividade do Plano` a
        WHERE {where}
        ORDER BY a.data IS NULL ASC, a.data ASC, a.name ASC
    """, params, as_dict=True)

    # ── Month grouping ─────────────────────────────────────────────────────
    MESES = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho",
             "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
    TODAY_KEY = _date.today().strftime("%Y-%m")

    def month_label(key):
        if key == "__nodate__":
            return "Sem Data Definida"
        y, m = key.split("-")
        return f"{MESES[int(m)-1]} {y}"

    groups = OrderedDict()
    for row in rows:
        key = str(row.data)[:7] if row.data else "__nodate__"
        groups.setdefault(key, []).append(row)

    sorted_keys = sorted(groups.keys(),
                         key=lambda k: ("\xff" if k == "__nodate__" else k))

    # ── Styles ─────────────────────────────────────────────────────────────
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _side(style="thin", color="D1D5DB"):
        return Side(style=style, color=color)

    def _border(all="thin"):
        s = _side(all)
        return Border(left=s, right=s, top=s, bottom=s)

    def _border_bottom(color="E5E7EB"):
        return Border(bottom=_side("thin", color))

    TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    HEADER_FONT  = Font(name="Calibri", bold=True, size=9,  color="374151")
    MONTH_FONT   = Font(name="Calibri", bold=True, size=9,  color="374151")
    BODY_FONT    = Font(name="Calibri", size=9,    color="1F2937")
    MUTED_FONT   = Font(name="Calibri", size=8,    color="9CA3AF", italic=True)
    NOTE_FONT    = Font(name="Calibri", size=8,    color="6B7280", italic=True)

    TITLE_FILL   = _fill("4F46E5")
    META_FILL    = _fill("EEF2FF")
    HEADER_FILL  = _fill("F1F5F9")
    MONTH_FILL   = _fill("E2E8F0")
    MONTH_CURR   = _fill("E0E7FF")
    ROW_ALT_FILL = _fill("F9FAFB")

    STATUS_FILLS = {
        "Pendente":     (_fill("F3F4F6"), Font(name="Calibri", size=9, color="6B7280")),
        "Em Progresso": (_fill("DBEAFE"), Font(name="Calibri", size=9, color="1D4ED8")),
        "Realizada":    (_fill("DCFCE7"), Font(name="Calibri", size=9, color="166534")),
        "Cancelada":    (_fill("FEE2E2"), Font(name="Calibri", size=9, color="991B1B")),
        "Adiada":       (_fill("FEF3C7"), Font(name="Calibri", size=9, color="92400E")),
    }

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
    left_mid = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Workbook ───────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano Anual"
    ws.sheet_view.showGridLines = False

    # Column widths: A(#) B(Actividade) C(Tipologia) D(Data) E(Orador) F(Local) G(Orçamento) H(Estado) I(Notas)
    col_widths = [4, 42, 18, 12, 24, 26, 14, 14, 38]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Title row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value        = f"📅  Plano Anual da Catequese — {ano_lectivo}"
    title_cell.font         = TITLE_FONT
    title_cell.fill         = TITLE_FILL
    title_cell.alignment    = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Meta row (filters applied) ─────────────────────────────────────────
    ws.row_dimensions[2].height = 16
    ws.merge_cells("A2:I2")
    meta_parts = [f"Exportado em {_date.today().strftime('%d/%m/%Y')}"]
    if estado:       meta_parts.append(f"Estado: {estado}")
    if tip_list:     meta_parts.append(f"Tipologia: {', '.join(tip_list)}")
    if month:        meta_parts.append(f"Mês: {month_label(month)}")
    meta_cell = ws["A2"]
    meta_cell.value     = "   ".join(meta_parts)
    meta_cell.font      = MUTED_FONT
    meta_cell.fill      = META_FILL
    meta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # ── Column headers ─────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 20
    headers = ["#", "Actividade", "Tipologia", "Data", "Orador / Responsável",
               "Local", "Orçamento", "Estado", "Notas de Execução"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = HEADER_FONT
        c.fill      = HEADER_FILL
        c.alignment = center if col == 1 else left_mid
        c.border    = Border(bottom=_side("medium", "CBD5E1"),
                             top=_side("thin", "E2E8F0"))

    # ── Data rows ──────────────────────────────────────────────────────────
    current_row = 4

    for key in sorted_keys:
        group_rows = groups[key]
        is_curr    = (key == TODAY_KEY)
        label      = month_label(key)

        # Month separator
        ws.row_dimensions[current_row].height = 18
        ws.merge_cells(f"A{current_row}:I{current_row}")
        mc = ws[f"A{current_row}"]
        mc.value     = f"  {label}  ·  {len(group_rows)} actividade{'s' if len(group_rows) != 1 else ''}"
        mc.font      = Font(name="Calibri", bold=True, size=9,
                            color="4338CA" if is_curr else "475569")
        mc.fill      = MONTH_CURR if is_curr else MONTH_FILL
        mc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col in range(1, 10):
            ws.cell(current_row, col).border = Border(
                top=_side("medium", "C7D2FE" if is_curr else "CBD5E1"),
                bottom=_side("thin",   "C7D2FE" if is_curr else "CBD5E1"),
            )
        current_row += 1

        # Activity rows
        for idx, row in enumerate(group_rows):
            ws.row_dimensions[current_row].height = 15
            fill = ROW_ALT_FILL if idx % 2 else None

            def _cell(col, value, font=BODY_FONT, align=left_mid, num_fmt=None):
                c = ws.cell(row=current_row, column=col, value=value)
                c.font = font
                c.alignment = align
                if fill: c.fill = fill
                c.border = _border_bottom()
                if num_fmt: c.number_format = num_fmt
                return c

            _cell(1, idx + 1, font=MUTED_FONT, align=center)
            _cell(2, row.actividade or "")
            _cell(3, row.tipologia  or "")
            date_val = str(row.data)[:10] if row.data else ""
            orig_val = str(row.data_original)[:10] if row.data_original else ""
            date_display = date_val
            if orig_val:
                date_display = f"{date_val}\n(orig: {orig_val})"
                ws.row_dimensions[current_row].height = 24
            _cell(4, date_display, align=Alignment(horizontal="center", vertical="center", wrap_text=bool(orig_val)))
            _cell(5, row.orador or "")
            _cell(6, row.local  or "", align=Alignment(horizontal="left", vertical="top", wrap_text=True))
            if row.orcamento:
                c = _cell(7, float(row.orcamento), align=Alignment(horizontal="right", vertical="center"))
                c.number_format = '#,##0.00 "MZN"'
            else:
                _cell(7, "")

            # Status — colour-coded
            st_fill, st_font = STATUS_FILLS.get(row.estado or "Pendente",
                                                STATUS_FILLS["Pendente"])
            sc = ws.cell(row=current_row, column=8, value=row.estado or "Pendente")
            sc.font      = st_font
            sc.fill      = st_fill
            sc.alignment = center
            sc.border    = _border_bottom()
            if fill: sc.fill = st_fill  # status colour always wins

            _cell(9, row.notas_execucao or "", font=NOTE_FONT,
                  align=Alignment(horizontal="left", vertical="top", wrap_text=True))

            current_row += 1

    # Freeze panes below header + col A
    ws.freeze_panes = "B4"

    # ── Totals footer ──────────────────────────────────────────────────────
    current_row += 1  # blank spacer
    ws.row_dimensions[current_row].height = 15
    ws.merge_cells(f"A{current_row}:G{current_row}")
    fc = ws[f"A{current_row}"]
    fc.value     = f"Total: {len(rows)} actividade{'s' if len(rows) != 1 else ''}"
    fc.font      = Font(name="Calibri", bold=True, size=9, color="6B7280")
    fc.alignment = Alignment(horizontal="right", vertical="center")
    fc.border    = Border(top=_side("medium", "E5E7EB"))
    for col in range(2, 9):
        ws.cell(current_row, col).border = Border(top=_side("medium", "E5E7EB"))

    # ── Serialize ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_ano = ano_lectivo.replace("/", "-")
    frappe.response["filename"]    = f"Plano_Anual_{safe_ano}.xlsx"
    frappe.response["filecontent"] = buf.read()
    frappe.response["type"]        = "download"


@frappe.whitelist()
def get_copy_preview(target_ano_lectivo):
    """Returns info about what would be copied — used to populate the confirmation modal."""
    _assert_coordenador()
    years = _sorted_anos()
    if target_ano_lectivo not in years:
        frappe.throw(_("Ano lectivo não encontrado"))
    idx = years.index(target_ano_lectivo)
    if idx == 0:
        frappe.throw(_("Não existe ano lectivo anterior"))
    prev_year = years[idx - 1]
    source_count = frappe.db.count("Actividade do Plano", {"ano_lectivo": prev_year})
    target_count = frappe.db.count("Actividade do Plano", {"ano_lectivo": target_ano_lectivo})
    return {"prev_year": prev_year, "source_count": source_count, "target_count": target_count}


@frappe.whitelist()
def copy_from_previous_year(target_ano_lectivo):
    """
    Copies all activities from the previous Ano Lectivo into target_ano_lectivo.
    - Shifts all dates forward by exactly 1 year (Feb 29 → Feb 28 in non-leap years)
    - Resets estado → Pendente
    - Clears notas_execucao and data_original
    - Preserves actividade, tipologia, orador, local, orcamento
    Returns the newly created rows (with tipologia details) for immediate UI update.
    """
    _assert_coordenador()
    import calendar as _cal
    from datetime import date as _date

    years = _sorted_anos()
    if target_ano_lectivo not in years:
        frappe.throw(_("Ano lectivo não encontrado"))
    idx = years.index(target_ano_lectivo)
    if idx == 0:
        frappe.throw(_("Não existe ano lectivo anterior para copiar"))
    prev_year = years[idx - 1]

    src_rows = frappe.db.sql("""
        SELECT actividade, tipologia, data, orador, local, orcamento
        FROM `tabActividade do Plano`
        WHERE ano_lectivo = %s
        ORDER BY data IS NULL ASC, data ASC, name ASC
    """, (prev_year,), as_dict=True)

    if not src_rows:
        frappe.throw(_("O ano anterior ({0}) não tem actividades para copiar").format(prev_year))

    def _shift(d):
        if not d:
            return None
        # d may come back as a datetime.date or string
        if not hasattr(d, "year"):
            try:
                from datetime import datetime
                d = datetime.strptime(str(d)[:10], "%Y-%m-%d").date()
            except Exception:
                return None
        new_year = d.year + 1
        last_day = _cal.monthrange(new_year, d.month)[1]
        return str(_date(new_year, d.month, min(d.day, last_day)))

    created_names = []
    for row in src_rows:
        doc = frappe.new_doc("Actividade do Plano")
        doc.actividade     = row.actividade
        doc.tipologia      = row.tipologia or None
        doc.estado         = "Pendente"
        doc.ano_lectivo    = target_ano_lectivo
        doc.data           = _shift(row.data)
        doc.orador         = row.orador or None
        doc.local          = row.local or None
        doc.orcamento      = row.orcamento or None
        doc.notas_execucao = None
        doc.insert(ignore_permissions=True)
        created_names.append(doc.name)

    frappe.db.commit()

    if not created_names:
        return {"copied": 0, "prev_year": prev_year, "rows": []}

    placeholders = ",".join(["%s"] * len(created_names))
    rows = frappe.db.sql(f"""
        SELECT a.name, a.actividade, a.data, a.data_original,
               a.orador, a.local, a.orcamento,
               a.tipologia, a.estado, a.notas_execucao, a.ano_lectivo,
               t.cor AS tipologia_cor, t.icone AS tipologia_icone
        FROM `tabActividade do Plano` a
        LEFT JOIN `tabTipologia Actividade` t ON t.name = a.tipologia
        WHERE a.name IN ({placeholders})
        ORDER BY a.data IS NULL ASC, a.data ASC, a.name ASC
    """, created_names, as_dict=True)

    return {"copied": len(created_names), "prev_year": prev_year, "rows": rows}


def _sorted_anos():
    """Return all Ano Lectivo names sorted ascending."""
    rows = frappe.db.sql(
        "SELECT name FROM `tabAno Lectivo` ORDER BY name ASC", as_dict=True
    )
    return [r.name for r in rows]


@frappe.whitelist()
def bulk_update_estado(names_json, estado):
    _assert_coordenador()
    allowed = {"Pendente", "Em Progresso", "Realizada", "Cancelada", "Adiada"}
    if estado not in allowed:
        frappe.throw(_("Estado inválido"))
    names = json.loads(names_json) if isinstance(names_json, str) else names_json
    if not names:
        frappe.throw(_("Nenhuma actividade seleccionada"))
    for name in names:
        if frappe.db.exists("Actividade do Plano", name):
            frappe.db.set_value("Actividade do Plano", name, "estado", estado)
    frappe.db.commit()
    return {"updated": len(names)}


@frappe.whitelist()
def bulk_delete(names_json):
    _assert_coordenador()
    names = json.loads(names_json) if isinstance(names_json, str) else names_json
    if not names:
        frappe.throw(_("Nenhuma actividade seleccionada"))
    deleted = 0
    for name in names:
        if frappe.db.exists("Actividade do Plano", name):
            frappe.delete_doc("Actividade do Plano", name, ignore_permissions=True)
            deleted += 1
    frappe.db.commit()
    return {"deleted": deleted}


@frappe.whitelist()
def bulk_move_month(names_json, new_month):
    """
    Move activities to a new month (YYYY-MM), preserving the day where possible.
    Saves data_original if this is the first date change.
    """
    _assert_coordenador()
    import re, calendar
    from datetime import date as _date
    if not re.match(r'^\d{4}-\d{2}$', new_month):
        frappe.throw(_("Formato de mês inválido"))
    names = json.loads(names_json) if isinstance(names_json, str) else names_json
    if not names:
        frappe.throw(_("Nenhuma actividade seleccionada"))

    ny, nm = map(int, new_month.split('-'))
    last_day = calendar.monthrange(ny, nm)[1]
    updated_rows = []

    for name in names:
        row = frappe.db.get_value(
            "Actividade do Plano", name,
            ["data", "data_original"], as_dict=True
        )
        if not row:
            continue
        if row.data:
            old_day = row.data.day if hasattr(row.data, 'day') else int(str(row.data)[8:10])
            new_day  = min(old_day, last_day)
            new_date = _date(ny, nm, new_day)
        else:
            new_date = _date(ny, nm, 1)

        update_vals = {"data": str(new_date)}
        new_data_original = None
        if row.data and not row.data_original:
            update_vals["data_original"] = str(row.data)
            new_data_original = str(row.data)

        frappe.db.set_value("Actividade do Plano", name, update_vals)
        updated_rows.append({
            "name": name,
            "data": str(new_date),
            "data_original": new_data_original or (str(row.data_original) if row.data_original else None),
        })

    frappe.db.commit()
    return {"updated": len(updated_rows), "rows": updated_rows}


@frappe.whitelist()
def reorder_actividades(ano_lectivo, ordered_names):
    """
    Persist drag-and-drop order within a month by updating a sort_order field.
    ordered_names is a JSON list of document names in the new order.
    """
    _assert_coordenador()
    names = json.loads(ordered_names) if isinstance(ordered_names, str) else ordered_names
    for idx, name in enumerate(names):
        frappe.db.set_value("Actividade do Plano", name, "idx", idx)
    frappe.db.commit()
    return {"success": True}
